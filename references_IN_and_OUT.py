import pandas as pd
import os
import sys
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import matplotlib.pyplot as plt
from matplotlib.patches import Patch  
import seaborn as sns
from helper_stuff import regions_to_color

# Get the current script's directory
script_directory = os.path.dirname(os.path.abspath(__file__))

# Get the parent directory
parent_directory = os.path.dirname(script_directory)

# Add the parent directory to the system path
sys.path.insert(0, parent_directory)

from find_ref_helper import clean_text2

def calculate_OUT_vs_total_edges(filename, df = None):
    
    # Read the xlsx file
    if filename != "":
        df = pd.read_excel(filename, engine='openpyxl', index_col=0)
    
    # Drop the "total" column
    df = df.drop('total', axis=1, errors='ignore')

    print(df)
    region_OUT_vs_total_dict = {}
    for index, row in df.iterrows():
        OUT_numerical_values = np.sum([value for col_name, value in row.items() if ((col_name != index) and col_name != "Middle East")])
        Total_numerical_values = np.sum([value for col_name, value in row.items() if ( col_name != "Middle East")])
        region_OUT_vs_total_dict[index] = OUT_numerical_values / Total_numerical_values

    print(region_OUT_vs_total_dict)
    return region_OUT_vs_total_dict

def plot_region_OUT_vs_total_edges(data):

    # Sort the data by values (region numbers)
    sorted_data = {k: v for k, v in sorted(data.items(), key=lambda item: item[1])}

    # Extract the sorted regions and their corresponding values
    regions = list(sorted_data.keys())
    values = list(sorted_data.values())

    # Get the colors for each region
    colors = [regions_to_color[region] for region in regions]

    # Plot the bar chart
    plt.figure(figsize=(10, 6))
    plt.bar(regions, values, color=colors)
    plt.xlabel('Community')
    plt.ylabel('Citations’ external out-degree')
    #plt.title('Ratio of community citation distribution: amount of cross-community citations divided by amount of total citations')
    plt.show()

# This function receives a dict where the key us the region and the value is another dict,
# were the key is teh regions (yes again) and the values are list of distinct authors list
# the function convert the lists of authors to the len of the lists 
def convert_distinct_author_dict_2_total(authors_list_per_region_dict):

    for region, inner_dict in authors_list_per_region_dict.items():
        inner_dict.pop('TargetRegion', None)
        inner_dict.pop("Middle East", None)

        # Normalize to percentage
        for inner_region in inner_dict:
            inner_dict[inner_region] = len(inner_dict[inner_region]) 
    
    return authors_list_per_region_dict

def plot_distinct_author_edges_accumulative_bar_chart(region_dict):

    region_dict.pop("SourceRegion", None)
    region_dict.pop("Middle East", None)

    # Generate a consistent color map for regions
    color_map = regions_to_color
    unique_regions = list(region_dict.keys())

    plt.figure(figsize=(10, 7)) 
    idx = -1

    for region, inner_dict in region_dict.items():
        bottom = 0
        idx += 1
        total_sum = sum(author_list_len for author_list_len in inner_dict.values())

        # Normalize to percentage
        for inner_region in inner_dict:
            inner_dict[inner_region] = ((inner_dict[inner_region]) / total_sum) * 100

        # Plot the bar
        for inner_region, value in inner_dict.items():
            #if inner_region != region:
                plt.bar(region, value, bottom=bottom, color=color_map.get(inner_region, "grey")) 
                bottom += value


    # Create a legend
    legend_patches = [Patch(color=color_map[region], label=region) for region in unique_regions]
    #item = Patch(color=color_map["Middle East"], label="Agreed Resources")
    #legend_patches.append(item)
    plt.legend(handles=legend_patches, title='Regions', bbox_to_anchor=(1.05, 1), loc='upper left')

    #plt.xticks(np.arange(len(sorted_names)), sorted_names, rotation=45, ha='right')  # Rotate names for better readability
    plt.title("Usage distribution of distinct authors across aegions")
    plt.ylabel('Authors from various communities')
    plt.tight_layout()
    plt.show()

def calculate_distinct_author_edges_RATIO_EXPECTED(data):
    data.pop("SourceRegion", None)
    region_distinct_authors_ratio_EXPECTED = {}

    for outer_region, inner_dict in data.items():
        inner_dict.pop("SourceRegion", None)
        full_value = 0
        external_value = 0
        for region, distinct_value in inner_dict.items():
            if (region != "Middle East"):
                full_value += distinct_value
            if (region != outer_region) and (region != "Middle East"):
                external_value += distinct_value
                
        # once finishing looping on the al the inners list of a region
        region_distinct_authors_ratio_EXPECTED[outer_region] = external_value/full_value
    
    region_distinct_authors_ratio_EXPECTED.pop("Middle East", None)

    return region_distinct_authors_ratio_EXPECTED

def plot_distinct_author_edges(data, Display = True, df = None):

    data.pop("SourceRegion", None)
    region_distinct_authors_ratio = {}

    for outer_region, inner_dict in data.items():
        full_list_value = []
        external_list_value = []
        inner_dict.pop("SourceRegion", None)
        for region, authors_list in inner_dict.items():
            if (region != "Middle East"):
                full_list_value.extend(authors_list)
            if (region != outer_region) and (region != "Middle East"):
                external_list_value.extend(authors_list)
        # once finishing looping on the al the inners list of a region
        total_external_edges = len(external_list_value)
        total_edges = len(full_list_value)
        region_distinct_authors_ratio[outer_region] = total_external_edges/total_edges
    
    region_distinct_authors_ratio.pop("Middle East", None)
    
    # Sort the data by values (region numbers)
    sorted_data = {k: v for k, v in sorted(region_distinct_authors_ratio.items(), key=lambda item: item[1])}

    # Extract the sorted regions and their corresponding values
    regions = list(sorted_data.keys())
    values = list(sorted_data.values())

    # Get the colors for each region
    colors = [regions_to_color[region] for region in regions]

    # Plot the bar chart
    if Display:
        plt.figure(figsize=(10, 6))
        plt.bar(regions, values, color=colors)
        plt.xlabel('Community')
        plt.ylabel('Resource diversity')
        plt.title("")# 'Ratio of regional distinct authors distribution: total vs. cross-region')
        plt.show()

    return region_distinct_authors_ratio

def build_ABS_referred_author_list(input_file, output_file):
    # Load the input .xlsx file as a pandas DataFrame
    df = pd.read_excel(input_file)

    # Create a new DataFrame which is grouped by 'Target' with the sum of 'Count'
    df_target_counts = df.groupby('Target')['Count'].sum().reset_index()

    # Sort the DataFrame in descending order by 'Count'
    df_target_counts = df_target_counts.sort_values(by='Count', ascending=False)

    # Create a dictionary where the keys are the 'Target' names and values are the counts
    ref_dict = dict(zip(df_target_counts['Target'], df_target_counts['Count']))

    # Write the DataFrame to a new .xlsx file
    df_target_counts.to_excel(output_file, index=False)

    # Print out the results
    for author, count in ref_dict.items():
        print(f'Author: {author}, Referred: {count} times')
        
    return ref_dict

def build_pivot_author_2_author_table(high_ref_file, edge_file, bio_file, ref_threshold, max_value_threshold, output_file):
    # Load the high reference file
    df_high_ref = pd.read_excel(high_ref_file)

    # Filter authors whose reference count is higher than the reference threshold
    highly_referred_authors = df_high_ref[df_high_ref['Count'] > ref_threshold]['Target'].tolist()

    # Load the edges file
    df_edges = pd.read_excel(edge_file)

    # Filter the DataFrame to only contain rows where Target is in highly_referred_authors
    df_edges = df_edges[df_edges['Target'].isin(highly_referred_authors)]

    # Load the biographies file
    df_bio = pd.read_excel(bio_file)

    # Merge biography data into edges data, including 'DATE OF BIRTH'
    df_edges = df_edges.merge(df_bio[['EN_NAME', 'REGION', 'DATE OF BIRTH']], how='left', left_on='Source', right_on='EN_NAME').rename(columns={'REGION': 'SourceRegion', 'DATE OF BIRTH': 'SourceDOB'})
    df_edges = df_edges.merge(df_bio[['EN_NAME', 'REGION', 'DATE OF BIRTH']], how='left', left_on='Target', right_on='EN_NAME').rename(columns={'REGION': 'TargetRegion', 'DATE OF BIRTH': 'TargetDOB'})

    # Sort DataFrame by 'Region' and 'DATE OF BIRTH' before creating pivot
    df_edges.sort_values(['SourceRegion', 'SourceDOB', 'Source', 'TargetRegion', 'TargetDOB', 'Target'], inplace=True)

    # Pivot the DataFrame to have 'Source' as index, 'Target' as columns and 'Count' as values, fill NaN for non-values
    pivot_df = df_edges.pivot_table(index=['SourceRegion', 'SourceDOB', 'Source'], columns=['TargetRegion', 'TargetDOB', 'Target'], values='Count', fill_value=np.nan)

    # Convert DataFrame to numeric
    pivot_df = pivot_df.apply(pd.to_numeric, errors='coerce')

    # Filter out rows where max value is less than the max_value_threshold
    pivot_df = pivot_df[pivot_df.max(axis=1) >= max_value_threshold]

    # Write the DataFrame to a new .xlsx file
    with pd.ExcelWriter(output_file) as writer:
        pivot_df.to_excel(writer, sheet_name='Full_Data')

        # Remove columns related to authors from the Middle East and save to another sheet
        pivot_df_no_me = pivot_df.loc[:,~pivot_df.columns.get_level_values(0).str.contains('Middle East')]
        pivot_df_no_me.to_excel(writer, sheet_name='Data_Excluding_ME')

    # Load workbook
    wb = load_workbook(output_file)
    ws1 = wb['Full_Data']
    ws2 = wb['Data_Excluding_ME']

    # Create fills
    dark_red_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    light_red_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
    regular_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Create font
    white_font = Font(color="FFFFFF")

    # Function to apply formatting
    def apply_formatting(ws):
        for row in ws.iter_rows(min_row=2, min_col=4, max_row=ws.max_row, max_col=ws.max_column):
            values = [cell.value for cell in row if isinstance(cell.value, int) or isinstance(cell.value, float)]
            if values:
                max_value = max(values)
                for cell in row:
                    if isinstance(cell.value, int) or isinstance(cell.value, float):
                        if cell.value >= (2/3 * max_value):
                            cell.fill = dark_red_fill
                            cell.font = white_font
                        elif cell.value <= (1/3 * max_value) and cell.value != 0:
                            cell.fill = light_red_fill
                        elif cell.value > 0:
                            cell.fill = regular_red_fill

    # Apply formatting to both worksheets
    apply_formatting(ws1)
    apply_formatting(ws2)

    # Save workbook
    wb.save(output_file)

    return pivot_df, pivot_df_no_me

def build_pivot_region_2_region_table__normalized_by_word_count(file_name, bio_file, word_count_file, output_file, apply_log = False):
    # Load the bio file and clean the 'NAME' column
    df_bio = pd.read_excel(bio_file, engine='openpyxl')
    df_bio['NAME'] = df_bio['NAME'].apply(lambda x: clean_text2(x)[0])

    # Load the word count file and clean the 'text_author' column
    df_word_count = pd.read_excel(word_count_file, engine='openpyxl')
    df_word_count['text_author'] = df_word_count['text_author'].apply(lambda x: clean_text2(x)[0])

    # Merge the two dataframes on the Hebrew name
    merged_df = pd.merge(df_bio, df_word_count, left_on='NAME', right_on='text_author')
    
    # Load the xlsx file
    df = pd.read_excel(file_name, header=None, engine='openpyxl')

    # Calculate total word count for each region
    total_word_count_by_region = merged_df.groupby('REGION')['Word Count'].sum()

    print(total_word_count_by_region)

    # Normalize the references by word count starting from the 3rd row and 3rd column
    '''
    row_len = df.shape[0]
    col_len = df.shape[1]
    for row in range(4, df.shape[0]):  # Start from the 3rd row
        for col in range(3, df.shape[1]):  # Start from the 3rd column
            cell_value = df.iat[row, col]
            if pd.notna(cell_value) and not isinstance(cell_value, str):
                author_name = df.iat[row, 2]  # Author names are in the 2nd row
                if author_name in merged_df['EN_NAME'].values:
                    word_count = merged_df[merged_df['EN_NAME'] == author_name]['Word Count'].iloc[0]
                    word_count = pd.to_numeric(word_count, errors='coerce')
                    cell_value = pd.to_numeric(cell_value, errors='coerce')

                    # Perform division if word_count is numeric and not zero
                    if pd.notna(word_count) and word_count != 0:
                        #if apply_log:
                        #    df.iat[row, col] = abs(math.log(cell_value / word_count))
                        #    df.iat[row, col] = abs(math.log10(cell_value / word_count))
                        #    df.iat[row, col] = abs(math.sqrt(cell_value / word_count))
                        #else:
                        df.iat[row, col] = cell_value / word_count
                    else:
                        df.iat[row, col] = 0
    '''

    # Find where the regions start in the first row and first column
    row_starts = [i for i in range(4, df.shape[0]) if pd.notnull(df.iloc[i, 0])]
    col_starts = [i for i in range(3, df.shape[1]) if pd.notnull(df.iloc[0, i])]

    # Append the last row and column index to include the last region
    row_starts.append(df.shape[0])
    col_starts.append(df.shape[1])

    # Get region names from the first row and first column
    regions_row = df.iloc[row_starts[:-1], 0].tolist()
    regions_col = df.iloc[0, col_starts[:-1]].tolist()


    #--------------------- Lets check for every AUTHOR how many references he has
    regions_authors_references_dict = {}
    authors_references_dict = {}
    region_row = ""
    for row_idx in range(row_starts[0],row_starts[-1]):
        # if its not nan it means we have a new region
        if pd.notna(df.iat[row_idx, 0]):
            # If we just finished to loop over all the authors (rows) of a region, 
            # and its not the first region, then, lets store it
            if region_row != "":
                if region_row !='Middle East':
                    regions_authors_references_dict[region_row] = authors_references_dict
                authors_references_dict = {}

            region_row = df.iat[row_idx, 0]

        #Skip SourceRegion and Middle East
        if region_row in ['SourceRegion', 'Middle East'] :
            continue
        
        author_name = df.iat[row_idx, 2]
        total_ref_count = 0
        external_ref_count = 0
        #now lets loop and sum all the column for the given author
        for column_idx in range(3, col_starts[-1]):
            if pd.notna(df.iat[0, column_idx]):
                region_col = df.iat[0, column_idx]
            #Skip SourceRegion and Middle East
            if region_col in ['SourceRegion', 'Middle East'] :
                continue   

            #extract the value
            value = df.iat[row_idx, column_idx]
            if pd.notna(value):
                total_ref_count += value
            else: continue
            # check if we are at a column of another region hence, its an external ref
            if region_row != region_col:
                external_ref_count += value
            else:
                pass
        
        # once finish to loop over all the columns
        # lets add the author to the authors dict:
        authors_references_dict[author_name] = [total_ref_count,external_ref_count]

    # add the last region
    regions_authors_references_dict[region_row] = authors_references_dict

    print(regions_authors_references_dict)
    
    #--------------------------------

    # Initialize result DataFrame
    result = pd.DataFrame(index=regions_row, columns=regions_col)

    for i in range(len(row_starts)-1):
        for j in range(len(col_starts)-1):
            # Define start and end index for rows and columns
            row_start, row_end = row_starts[i], row_starts[i+1]
            col_start, col_end = col_starts[j], col_starts[j+1]

            # Slice DataFrame to get the current 'Big square'
            big_square = df.iloc[row_start:row_end, col_start:col_end]
            
            # Sum up all the elements in the 'Big square' and put the result in the new DataFrame
            result.iloc[i, j] = big_square.sum().sum()

    print(result)

    # Filter out 'Middle East' region
    result = result.drop('Middle East', errors='ignore') # Remove row
    result = result.drop('Middle East', axis=1, errors='ignore') # Remove column

    total_full_network_citation_ = result.sum().sum()

    # divide the number of references by the number of words
    result = result.div(total_word_count_by_region, axis=0)

    print(result)

    
    total_word_count = 0
    for region, word_count in total_word_count_by_region.items():
        print(f"{region}: {word_count}")
        if region != "Middle East":
            total_word_count += word_count

    print("total_full_network_citation: ", total_full_network_citation_)
    print("total_word_count: ", total_word_count)
    full_Text_citation_rate = total_full_network_citation_/total_word_count
    print(" Full network text citation rate")
    full_Text_citation_rate = full_Text_citation_rate 
    print(full_Text_citation_rate )

    # Loop over each row by its index
    for row_index in result.index:
        if row_index == "Middle East":
            continue
        # Sum all values in the row except the value where column name equals the row index
        total = result.loc[row_index].drop(row_index).sum()
        # Assign the total to a new column 'total'
        result.loc[row_index, 'Text citation rate'] = total

    # Filter out 'Middle East' region
    #result = result.drop('Middle East', errors='ignore') # Remove row
    #result = result.drop('Middle East', axis=1, errors='ignore') # Remove column

    result.index.name = "REGION"
    print(result)

    # Save result DataFrame to a new Excel file
    result.to_excel(output_file, engine='openpyxl')

    result.loc['Total'] = full_Text_citation_rate
    return result, regions_authors_references_dict

def build_pivot_region_2_region_table_by_num_of_edges(file_name, bio_file, word_count_file, output_file, normalize = False):
    # Load the bio file and clean the 'NAME' column
    df_bio = pd.read_excel(bio_file, engine='openpyxl')
    df_bio['NAME'] = df_bio['NAME'].apply(lambda x: clean_text2(x)[0])

    # Load the word count file and clean the 'text_author' column
    df_word_count = pd.read_excel(word_count_file, engine='openpyxl')
    df_word_count['text_author'] = df_word_count['text_author'].apply(lambda x: clean_text2(x)[0])

    # Merge the two dataframes on the Hebrew name
    _ = pd.merge(df_bio, df_word_count, left_on='NAME', right_on='text_author')
    
    # Load the xlsx file
    df = pd.read_excel(file_name, header=None, engine='openpyxl')

    # converting the values in the table to 1 or 0 (as we don't care about the value, we just want to know if there is a value or not)
    for row in range(4, df.shape[0]):  # Start from the 3rd row
        for col in range(3, df.shape[1]):  # Start from the 3rd column
            cell_value = df.iat[row, col]
            if pd.notna(cell_value) and not isinstance(cell_value, str):
                df.iat[row, col] = 1


    # Find where the regions start in the first row and first column
    row_starts = [i for i in range(4, df.shape[0]) if pd.notnull(df.iloc[i, 0])]
    col_starts = [i for i in range(3, df.shape[1]) if pd.notnull(df.iloc[0, i])]

    # Append the last row and column index to include the last region
    row_starts.append(df.shape[0])
    col_starts.append(df.shape[1])

    # Get region names from the first row and first column
    regions_row = df.iloc[row_starts[:-1], 0].tolist()
    regions_col = df.iloc[0, col_starts[:-1]].tolist()

    # Get the list of authors from every region
    _, lengths_dict = get_corresponding_values_and_lengths(df)

    # Initialize result DataFrame
    Big_square_df = pd.DataFrame(index=regions_row, columns=regions_col)

    for i in range(len(row_starts)-1):
        for j in range(len(col_starts)-1):
            # Define start and end index for rows and columns
            row_start, row_end = row_starts[i], row_starts[i+1]
            col_start, col_end = col_starts[j], col_starts[j+1]

            # Slice DataFrame to get the current 'Big square'
            big_square = df.iloc[row_start:row_end, col_start:col_end]
            
            # Sum up all the elements in the 'Big square' and put the result in the new DataFrame
            Big_square_df.iloc[i, j] = big_square.sum().sum()


    # Filter out 'Middle East' region
    Big_square_df = Big_square_df.drop('Middle East', errors='ignore') # Remove row
    Big_square_df = Big_square_df.drop('Middle East', axis=1, errors='ignore') # Remove column

    #full_total_df = full_total_df.drop('Middle East', errors='ignore') # Remove row
    #full_total_df = full_total_df.drop('Middle East', axis=1, errors='ignore') # Remove column


    print(Big_square_df)
    # Normalize Big_square_ .0df based on lengths_dict
    if normalize == True:
        for idx in Big_square_df.index:  # Iterate over index
            if idx in lengths_dict:
                length = lengths_dict[idx]
                Big_square_df.loc[idx] = Big_square_df.loc[idx].apply(lambda x: x / length if pd.notna(x) else x)
    print(lengths_dict)
    print(Big_square_df)

    full_total_df =  pd.DataFrame(index=regions_row)
    # Loop over each row by its index
    for row_index in Big_square_df.index:
        # Sum all values in the row
        total_full = Big_square_df.loc[row_index].sum()
        full_total_df.loc[row_index, 'Author citation rate'] = total_full

        # Sum all values in the row except the value where column name equals the row index
        total = Big_square_df.loc[row_index].drop(row_index).sum()
        # Assign the total to a new column 'total'
        Big_square_df.loc[row_index, 'total'] = total


    '''
    # Filter out 'Middle East' region
    Big_square_df = Big_square_df.drop('Middle East', errors='ignore') # Remove row
    Big_square_df = Big_square_df.drop('Middle East', axis=1, errors='ignore') # Remove column

    full_total_df = full_total_df.drop('Middle East', errors='ignore') # Remove row
    full_total_df = full_total_df.drop('Middle East', axis=1, errors='ignore') # Remove column
    '''

    print(Big_square_df)
    # Save result DataFrame to a new Excel file
    Big_square_df.to_excel(output_file, engine='openpyxl')

    full_total_df.index.name = "REGION"

    return full_total_df
    
def build_pivot_region_2_region_table_edges_not_normalized(file_name, bio_file, word_count_file, output_file, normalize = False):
    # Load the bio file and clean the 'NAME' column
    df_bio = pd.read_excel(bio_file, engine='openpyxl')
    df_bio['NAME'] = df_bio['NAME'].apply(lambda x: clean_text2(x)[0])

    # Load the word count file and clean the 'text_author' column
    df_word_count = pd.read_excel(word_count_file, engine='openpyxl')
    df_word_count['text_author'] = df_word_count['text_author'].apply(lambda x: clean_text2(x)[0])

    # Merge the two dataframes on the Hebrew name
    merged_df = pd.merge(df_bio, df_word_count, left_on='NAME', right_on='text_author')
    
    # Load the xlsx file
    df = pd.read_excel(file_name, header=None, engine='openpyxl')

    # converting the values in the table to 1 or 0 (as we don't care about the value, we just want to know if there is a value or not)
    row_len = df.shape[0]
    col_len = df.shape[1]
    for row in range(4, df.shape[0]):  # Start from the 3rd row
        for col in range(3, df.shape[1]):  # Start from the 3rd column
            cell_value = df.iat[row, col]
            if pd.notna(cell_value) and not isinstance(cell_value, str):
                df.iat[row, col] = 1


    # Find where the regions start in the first row and first column
    row_starts = [i for i in range(4, df.shape[0]) if pd.notnull(df.iloc[i, 0])]
    col_starts = [i for i in range(3, df.shape[1]) if pd.notnull(df.iloc[0, i])]

    # Append the last row and column index to include the last region
    row_starts.append(df.shape[0])
    col_starts.append(df.shape[1])

    # Get region names from the first row and first column
    regions_row = df.iloc[row_starts[:-1], 0].tolist()
    regions_col = df.iloc[0, col_starts[:-1]].tolist()

    # Get the list of authors from every region
    corresponding_values_dict, lengths_dict = get_corresponding_values_and_lengths(df)

    # Initialize result DataFrame
    Big_square_df = pd.DataFrame(index=regions_row, columns=regions_col)

    for i in range(len(row_starts)-1):
        for j in range(len(col_starts)-1):
            # Define start and end index for rows and columns
            row_start, row_end = row_starts[i], row_starts[i+1]
            col_start, col_end = col_starts[j], col_starts[j+1]

            # Slice DataFrame to get the current 'Big square'
            big_square = df.iloc[row_start:row_end, col_start:col_end]
            
            # Sum up all the elements in the 'Big square' and put the result in the new DataFrame
            Big_square_df.iloc[i, j] = big_square.sum().sum()

    
    print(Big_square_df)
    Big_square_df = Big_square_df.drop('Middle East', errors='ignore') # Remove row
    Big_square_df = Big_square_df.drop('Middle East', axis=1, errors='ignore') # Remove column
    #print("=====================pivote table num of edges=================")
    print(Big_square_df)
    #print("===================== num of edge per region=================")
    #print(Big_square_df.sum(axis=1))

    sum_of_edges_per_region = Big_square_df.sum(axis=1)
    # Normalize Big_square_ .0df based on lengths_dict
    if normalize == True:
        for idx in Big_square_df.index:  # Iterate over index
            if idx in lengths_dict:
                length = lengths_dict[idx]
                Big_square_df.loc[idx] = Big_square_df.loc[idx].apply(lambda x: x / length if pd.notna(x) else x)
    print(lengths_dict)
    print(Big_square_df)

    # Loop over each row by its index
    for row_index in Big_square_df.index:
        # Sum all values in the row except the value where column name equals the row index
        total = Big_square_df.loc[row_index].drop(row_index).sum()
        # Assign the total to a new column 'total'
        Big_square_df.loc[row_index, 'total'] = total

    '''
    # Filter out 'Middle East' region
    Big_square_df = Big_square_df.drop('Middle East', errors='ignore') # Remove row
    Big_square_df = Big_square_df.drop('Middle East', axis=1, errors='ignore') # Remove column
    '''

    print(Big_square_df)
    # Save result DataFrame to a new Excel file
    Big_square_df.to_excel(output_file, engine='openpyxl')
    
    return sum_of_edges_per_region

def build_distinct_author_edges_per_region(file_name, bio_file, word_count_file, output_file, output_file_detailed_regions, normalize=False):
    # Load the bio file and clean the 'NAME' column
    df_bio = pd.read_excel(bio_file, engine='openpyxl')
    df_bio['NAME'] = df_bio['NAME'].apply(lambda x: clean_text2(x)[0])

    # Load the word count file and clean the 'text_author' column
    df_word_count = pd.read_excel(word_count_file, engine='openpyxl')
    df_word_count['text_author'] = df_word_count['text_author'].apply(lambda x: clean_text2(x)[0])

    # Merge the two dataframes on the Hebrew name
    merged_df = pd.merge(df_bio, df_word_count, left_on='NAME', right_on='text_author')

    # Load the xlsx file
    df = pd.read_excel(file_name, header=None, engine='openpyxl')

    # converting the values in the table to 1 or 0 (as we don't care about the value, we just want to know if there is a value or not)
    row_len = df.shape[0]
    col_len = df.shape[1]
    for row in range(4, df.shape[0]):  # Start from the 3rd row
        for col in range(3, df.shape[1]):  # Start from the 3rd column
            cell_value = df.iat[row, col]
            if pd.notna(cell_value) and not isinstance(cell_value, str):
                df.iat[row, col] = 1

    # Extract region names and corresponding rows
    region_rows = {}
    current_region = None
    for i in range(df.shape[0]):
        region_name = df.iat[i, 0]
        if pd.notnull(region_name):
            if current_region is not None:
                # Set the end row for the previous region
                region_rows[current_region].append(i - 1)
            current_region = region_name
            # Start a new region range
            region_rows[current_region] = [i]
    # Set the end row for the last region
    if current_region is not None:
        region_rows[current_region].append(df.shape[0] - 1)

    # Create a dictionary to hold region to authors mapping
    region_to_authors = {}

    # Iterate through each region and their corresponding rows
    for region, rows in region_rows.items():
        referred_authors = set()
        for row in rows:
            for col in range(3, df.shape[1]):  # Assuming the authors' names start from the 3rd column
                author = df.iat[2, col]  # Get the author's name from the second row
                if pd.notna(df.iat[row, col]) and pd.notna(author):
                    referred_authors.add(author)
        region_to_authors[region] = list(referred_authors)

    # Output the dictionary
    print(region_to_authors)

    # Convert the dictionary to a DataFrame for easy export
    max_length = max(len(authors) for authors in region_to_authors.values())
    # Create a DataFrame with an appropriate shape
    author_df = pd.DataFrame.from_dict(region_to_authors, orient='index')
    # Ensure the DataFrame has enough columns
    author_df = author_df.reindex(columns=range(max_length))

    # Saving the result to the output file
    author_df.to_excel(output_file, engine='openpyxl')
    
    #----------------------------------------------------------------------

    # Extract region names and corresponding rows and columns
    region_rows = {}
    region_cols = {}
    current_region_row = None
    current_region_col = None
    for i in range(df.shape[0]):  # Rows
        region_name_row = df.iat[i, 0]
        if pd.notnull(region_name_row):
            if current_region_row is not None:
                region_rows[current_region_row].append(i - 1)
            current_region_row = region_name_row
            region_rows[current_region_row] = [i]
    if current_region_row is not None:
        region_rows[current_region_row].append(df.shape[0] - 1)

    for j in range(df.shape[1]):  # Columns
        region_name_col = df.iat[0, j]
        if pd.notnull(region_name_col):
            if current_region_col is not None:
                region_cols[current_region_col].append(j - 1)
            current_region_col = region_name_col
            region_cols[current_region_col] = [j]
    if current_region_col is not None:
        region_cols[current_region_col].append(df.shape[1] - 1)

    # Create a nested dictionary to hold region to authors mapping
    region_to_authors_detailed = {}

    # Iterate through each region and their corresponding rows and columns
    for outer_region, outer_rows in region_rows.items():
        inner_dict = {}
        for inner_region, inner_cols in region_cols.items():
            referred_authors = set()
            for row in range(outer_rows[0], outer_rows[1] + 1):
                for col in range(inner_cols[0], inner_cols[1] + 1):
                    author = df.iat[2, col]  # Assuming authors' names start from the 3rd row
                    if pd.notna(df.iat[row, col]) and pd.notna(author):
                        referred_authors.add(author)
            inner_dict[inner_region] = list(referred_authors)
        region_to_authors_detailed[outer_region] = inner_dict

    # find the value of the last column
    last_key = list(region_cols.keys())[-1]
    last_column_idx = region_cols[last_key][1]
    regions_authors_references_dict = {}
    authors_references_dict = {}
    for region, outer_rows in region_rows.items():
        if region in ['SourceRegion', 'Middle East'] :
            continue
        # running over all the rows of teh outer_region
        authors_references_dict = {}
        for row_idx in range(outer_rows[0],outer_rows[1]+1):
            author_name = df.iat[row_idx, 2]
            total_ref_count = 0
            external_ref_count = 0
            # loop over all the columns is that row_idx
            for column_idx in range(3,last_column_idx + 1):
                
                # Since in our xls file the column are merged then the first cell of the column will have a value, 
                # and teh rest will be nan, hence we keep the previous value, and we change only when we have a real new value which is not nan
                if pd.notna(df.iat[0, column_idx]):
                    referred_region = df.iat[0, column_idx]
                # We skip the 'Middle East'
                if referred_region == 'Middle East':
                    continue
                #extract the value
                value = df.iat[row_idx, column_idx]
                if pd.notna(value):
                    total_ref_count += value
                else: continue
                # check if we are at a column of another region hence, its an external ref
                if region != referred_region:
                    external_ref_count += value
                else:
                    pass
            # once finish to loop over all the columns:
            authors_references_dict[author_name] = [total_ref_count,external_ref_count]

        #once authors dict is ready we can assign it to its region            
        regions_authors_references_dict[region] = authors_references_dict
    

    print(regions_authors_references_dict)

    # Output the nested dictionary
    print(region_to_authors_detailed)
    with pd.ExcelWriter(output_file_detailed_regions, engine='openpyxl') as writer:
        for region, authors_dict in region_to_authors_detailed.items():
            # Convert each inner dictionary to a DataFrame
            df = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in authors_dict.items()]))
            # Save the DataFrame to a sheet named after the region
            df.to_excel(writer, sheet_name=region)

    return region_to_authors, region_to_authors_detailed, regions_authors_references_dict

def get_corresponding_values_and_lengths(df):
    merged_cell_ranges = get_merged_cell_ranges(df)
    corresponding_values_dict = {}
    lengths_dict = {}
    for start, end in merged_cell_ranges:
        merged_cell_value = df.iloc[start, 0]  # Get the merged cell value
        values = df.iloc[start:end, 2].tolist()  # Get values from the third column
        corresponding_values_dict[merged_cell_value] = values
        lengths_dict[merged_cell_value] = len(values)  # Store the length of the list
    return corresponding_values_dict, lengths_dict

# Function to get the ranges of merged cells in the first column
def get_merged_cell_ranges(df):
    merged_ranges = []
    start = 4
    for i in range(5, len(df)):
        # If the cell is not merged (i.e., it has a value), then the previous range is ended
        if pd.notna(df.iloc[i, 0]):
            merged_ranges.append((start, i))  # End of the previous merged range
            start = i
    # Add the last range
    merged_ranges.append((start, len(df)))
    return merged_ranges

def build_pivot_region_2_region_table(file_name, output_file):
    # Load the xlsx file
    df = pd.read_excel(file_name, header=None, engine='openpyxl')

    # Find where the regions start in the first row and first column
    row_starts = [i for i in range(4, df.shape[0]) if pd.notnull(df.iloc[i, 0])]
    col_starts = [i for i in range(3, df.shape[1]) if pd.notnull(df.iloc[0, i])]

    # Append the last row and column index to include the last region
    row_starts.append(df.shape[0])
    col_starts.append(df.shape[1])

    # Get region names from the first row and first column
    regions_row = df.iloc[row_starts[:-1], 0].tolist()
    regions_col = df.iloc[0, col_starts[:-1]].tolist()

    # Initialize result DataFrame
    result = pd.DataFrame(index=regions_row, columns=regions_col)

    for i in range(len(row_starts)-1):
        for j in range(len(col_starts)-1):
            # Define start and end index for rows and columns
            row_start, row_end = row_starts[i], row_starts[i+1]
            col_start, col_end = col_starts[j], col_starts[j+1]

            # Slice DataFrame to get the current 'Big square'
            big_square = df.iloc[row_start:row_end, col_start:col_end]
            
            # Sum up all the elements in the 'Big square' and put the result in the new DataFrame
            result.iloc[i, j] = big_square.sum().sum()

    # Save result DataFrame to a new Excel file
    result.to_excel(output_file, engine='openpyxl')

def plot_region_2_region_table(file_name, color, title = "", percentage_normalization = False, df = None):
    
    
    if df is None:
        df = pd.read_excel(file_name, engine='openpyxl', index_col=0)

    print(df)
    
    # Filter out 'Middle East' region
    df = df.drop('Middle East', errors='ignore') # Remove row
    df = df.drop('Middle East', axis=1, errors='ignore') # Remove column
    df = df.drop('total', axis=1, errors='ignore') # Remove column

    '''
    df.rename(columns={'Middle East': 'Agreed Resources'}, inplace=True)

    # Get a list of all column names, except the one you want to move
    column_order = [col for col in df.columns if col != 'Agreed Resources']

    # Add the column you want to move to the end of the list
    column_order.append('Agreed Resources')

    # Reorder the columns using the new column order
    df = df[column_order]
    '''
    # Normalize each row by its max value
    if percentage_normalization == False:
        df_normalized = df.div(df.max(axis=1), axis=0)
    else:
        df_normalized = df.div(df.sum(axis=1), axis=0) * 100

    # Create a figure and axis
    fig, ax = plt.subplots(figsize=(10, 10))
    
    # Create a heatmap
    VMAX = None
    if percentage_normalization == False:
        VMAX = 1
    else:
        VMAX = df_normalized.max().max()
    sns.heatmap(df_normalized, cmap=color, linewidths=.5, annot=True, fmt=".2f", ax=ax, vmin=0, vmax=VMAX)
    
    # Position the x-axis labels on the top
    ax.xaxis.tick_top()
    ax.xaxis.set_label_position('top')

    # Improve the visibility of the labels by rotating them
    plt.xticks(rotation=45)
    plt.yticks(rotation=0)

    # Add title to the chart
    fig.suptitle(title)

    # Show the plot
    plt.show()


